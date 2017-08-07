#!/usr/bin/env python

#
# EAlGIS loader: Australian Census 2011; Data Pack 1
#

import re
import os
import glob
import os.path
import openpyxl
import sqlalchemy

from .ealgis.loaders import ZipAccess, ShapeLoader, RewrittenCSV, CSVLoader
from .ealgis.util import alistdir, make_logger
from .ealgis.db import EalLoader
from .ealgis.util import cmdrun
from sqlalchemy.schema import CreateSchema


logger = make_logger(__name__)


def go(loader, tmpdir):
    census_dir = '/app/data/2011 Datapacks BCP_IP_TSP_PEP_ECP_WPP_ERP_Release 3'
    release = '3'
    schema_name = "aus_census_2011"

    census_division_table = {}
    geo_gid_mapping = {}

    ###
    # unfortunately straight joins shape attribute to census CSV attribute don't work nicely;
    # after the shape load some of the attributes end up as char(9292) or some awful thing.
    # so we do a mapping in this code to our internal gid column; which has the nice property
    # of making things fast and obvious in production
    ###

    def mapper():
        cls = loader.get_table_class('sa1_2011_aust')
        loader.session.query(sqlalchemy.cast(cls.sa1_7digit, sqlalchemy.Integer()))[:10]

    def load_shapes():
        logger.debug("load shapefiles")
        new_tables = []

        def shapefiles():
            def shape_and_proj(g):
                for f in g:
                    shape_name = os.path.basename(f)
                    proj = shape_name.split('_')[1]
                    yield f, proj
            # favour the POW shapes over the others; release 3 eccentricity
            projs_provided = set()
            for fname, proj in shape_and_proj(glob.glob(os.path.join(census_dir, "Digital Boundaries/*_POW_shape.zip"))):
                projs_provided.add(proj)
                yield fname
            for fname, proj in shape_and_proj(glob.glob(os.path.join(census_dir, "Digital Boundaries/*_shape.zip"))):
                if proj not in projs_provided:
                    yield fname

        for fname in shapefiles():
            with ZipAccess(None, tmpdir, fname) as z:
                for shpfile in z.glob("*.shp"):
                    before = set(loader.get_table_names())
                    instance = ShapeLoader(shpfile, 4283)
                    instance.load(loader)
                    new = list(set(loader.get_table_names()) - before)
                    assert(len(new) == 1)
                    new_tables.append(new[0])

        logger.info("loaded shapefile OK")

        logger.info("creating shape indexes")
        # create column indexes on shape linkage
        loader.session.commit()
        for census_division in shp_linkage:
            pfx = "%s_2011" % (census_division)
            table = [t for t in new_tables if t.startswith(pfx)][0]
            census_division_table[census_division] = table
            info = loader.get_table(table)
            col, _, descr = shp_linkage[census_division]
            loader.set_table_metadata(table, {'description': descr})
            idx = sqlalchemy.Index("%s_%s_idx" % (table, col), info.columns[col], unique=True)
            idx.create(loader.engine)
            logger.debug(repr(idx))

        # create geo_column -> gid mapping
        logger.info("creating gid mapping tables")
        for census_division in shp_linkage:
            geo_table = census_division_table[census_division]
            geo_column, geo_cast_required, _ = shp_linkage[census_division]
            geo_cls = loader.get_table_class(geo_table)
            geo_attr = getattr(geo_cls, geo_column)
            if geo_cast_required is not None:
                inner_col = sqlalchemy.cast(geo_attr, geo_cast_required)
            else:
                inner_col = geo_attr
            logger.debug(repr([geo_table, geo_column, inner_col]))
            lookup = {}
            for gid, match in loader.session.query(geo_cls.gid, inner_col).all():
                lookup[str(match)] = gid
            geo_gid_mapping[census_division] = lookup

    data_tables = []

    def load_datapacks(packname):
        def get_csv_files():
            files = []
            for geography in alistdir(d):
                g = os.path.join(geography, "*.csv")
                csv_files = glob.glob(g)
                if len(csv_files) == 0:
                    g = os.path.join(geography, "AUST", "*.csv")
                    csv_files = glob.glob(g)
                if len(csv_files) == 0:
                    raise Exception("can't find CSV files for `%s'" % geography)
                files += csv_files
            return files

        d = os.path.join(census_dir, packname, "Sequential Number Descriptor")
        csv_files = get_csv_files()
        table_re = re.compile(r'^2011Census_(.*)_sequential.csv$')
        linkage_pending = []

        for i, csv_path in enumerate(csv_files):
            if i > 3:
                break
            logger.info("[%d/%d] %s: %s" % (i + 1, len(csv_files), packname, os.path.basename(csv_path)))
            table_name = table_re.match(os.path.split(csv_path)[-1]).groups()[0].lower()
            data_tables.append(table_name)
            decoded = table_name.split('_')
            census_table, census_country = table_name[0], table_name[1]
            if len(decoded) == 3:
                census_division = decoded[2]
            else:
                census_division = None

            gid_match = None

            if census_division is not None:
                def make_match_fn():
                    lookup = geo_gid_mapping[census_division]

                    def _matcher(line, row):
                        if line == 0:
                            # rewrite the header
                            return ['gid'] + row
                        else:
                            return [str(lookup[row[0]])] + row
                    return _matcher
                gid_match = make_match_fn()

            # normalise the CSV file by reading it in and writing it out again,
            # Postgres is quite pedantic. we also want to add an additional column to it
            with RewrittenCSV(tmpdir, csv_path, gid_match) as norm:
                instance = CSVLoader(table_name, norm.get(), pkey_column=0)
                table_info = instance.load(loader)
                if table_info is not None and census_division is not None:
                    linkage_pending.append((table_name, table_info, census_division))

        # done as another pass to avoid having to re-run the reflection of the entire
        # database for every CSV file loaded (can be thousands)
        for attr_table, table_info, census_division in linkage_pending:
            geo_table = census_division_table[census_division]
            geo_column, _, _ = shp_linkage[census_division]
            loader.add_geolinkage(
                geo_table, "gid",
                attr_table, "gid")

    def load_metadata(*fnames):
        def load_workbook(fname):
            logger.info("parsing metadata: %s" % (fname))
            wb = openpyxl.load_workbook(fname, use_iterators=True)

            def sheet_data(sheet):
                return ([t.internal_value for t in r] for r in sheet.iter_rows())

            def skip(it, n):
                for i in range(n):
                    next(it)

            sheet_iter = sheet_data(wb.worksheets[0])
            skip(sheet_iter, 3)
            for row in sheet_iter:
                name = row[0]
                if not name:
                    continue
                name = name.lower()
                table_meta[name] = {'type': row[1], 'kind': row[2]}

            sheet_iter = sheet_data(wb.worksheets[1])
            skip(sheet_iter, 4)
            for row in sheet_iter:
                name = row[0]
                if not name:
                    continue
                name = name.lower()
                short_name, long_name, datapack_file, profile_table, column_heading = row[1:6]
                datapack_file = datapack_file.lower()
                if datapack_file not in col_meta:
                    col_meta[datapack_file] = []
                col_meta[datapack_file].append((name, {'type': row[2], 'kind': row[5]}))
            del wb

        table_meta = {}
        col_meta = {}
        for fname in fnames:
            load_workbook(os.path.join(census_dir, os.path.join('Metadata/', fname)))

        for table_name in data_tables:
            datapack_file = table_name.split('_', 1)[0].lower()
            m = re.match('^([A-Za-z]+[0-9]+)([a-z]+)?$', datapack_file)
            table_number = m.groups()[0]
            meta = table_meta[table_number]
            columns = col_meta[datapack_file]
            loader.set_table_metadata(table_name, meta)
            loader.register_columns(table_name, columns)
    
    loader.set_metadata(
        name="ABS Census 2011",
        description="Shapes")

    logger.info("created metadata record - version %s in `ealgis_metadata`" % (first_version.version))

    load_shapes()

    load_datapacks("2011 Aboriginal and Torres Strait Islander Peoples Profile Release %s" % release)
    load_datapacks("2011 Basic Community Profile Release %s" % release)
    load_datapacks("2011 Expanded Community Profile Release %s" % release)
    load_datapacks("2011 Place of Enumeration Profile Release %s" % release)
    load_datapacks("2011 Time Series Profile Release %s" % release)
    load_datapacks("2011 Working Population Profile Release %s" % release)

    load_metadata(
        "Metadata_2011_BCP_DataPack.xlsx",
        "Metadata_2011_IP_DataPack.xlsx",
        "Metadata_2011_PEP_DataPack.xlsx",
        "Metadata_2011_TSP_DataPack.xlsx",
        "Metadata_2011_WPP_DataPack.xlsx",
        "Metadata_2011_XCP_DataPack.xlsx")

    logger.info("create schema %s" % schema_name)
    loader.engine.execute(CreateSchema(schema_name))

    logger.info("move tables to standalone schema")
    ealgis_tables = ["user", "setting", "geometry_touches", "map_definition", "geometry_intersection", "geometry_relation", "spatial_ref_sys"]
    for table_name in loader.get_table_names():
        if table_name not in ealgis_tables:
            try:
                loader.engine.execute('ALTER TABLE %s SET SCHEMA %s;' % (table_name, schema_name))
                loader.session.commit()
                logger.info(table_name)
            except sqlalchemy.exc.ProgrammingError as e:
                logger.info("couldn't change schema for table: %s (%s)" % (table_name, e))

    logger.info("dumping database")
    os.environ['PGPASSWORD'] = loader.dbpassword()
    shp_cmd = ["pg_dump", str(loader.engineurl()), "--schema=%s" % schema_name, "--format=c", "--file=/app/tmp/%s" % schema_name]

    stdout, stderr, code = cmdrun(shp_cmd)
    if code != 0:
        raise Exception("database dump with pg_dump failed: %s." % stderr)
    else:
        logger.info("successfully dumped database to /app/tmp/%s" % schema_name)
        logger.info("load with: pg_restore --username=user --dbname=db /path/to/%s" % schema_name)
        logger.info("then run VACUUM ANALYZE;")


if __name__ == '__main__':
    loader = EalLoader(
        "aucensus2011",
        mandatory_srids=[3112, 3857])
    tmpdir = "/app/tmp"
    go(loader, tmpdir)
    logger.info("OK")